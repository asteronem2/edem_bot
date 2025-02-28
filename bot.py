import asyncio
import datetime
import json
import logging
import os
import re
import time
import traceback
from datetime import timezone, timedelta
from typing import Union, Literal, List
import json

import aiogram
import httpx
from aiogram import Bot, Dispatcher, types as tg_types, F, exceptions
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Filter
from aiogram.types import InlineKeyboardButton as IButton, InlineKeyboardMarkup, FSInputFile, Message, CallbackQuery
from mako.compat import win32
from openpyxl import Workbook

from core import UserCore, UserMessageCore, BotMessageCore, PaymentCore
import messages
from messages import MsgModel, config

logging.basicConfig(level=logging.WARNING, filename='logs.log', filemode='w',
                    format='%(asctime)s - [%(levelname)s] - %(message)s')

bot = Bot(token=config.BOT_TOKEN, default=DefaultBotProperties(parse_mode='html'))
dp = Dispatcher()


async def log(message: str, level: str = 'error'):
    if level == 'error':
        logging.error(message)
        try:
            await bot.send_document(892097042, FSInputFile('logs.log'), disable_notification=True)
            await bot.send_message(892097042, message, disable_notification=True)
            with open("logs.log", "w"):
                pass
        except Exception as err:
            err_str = traceback.format_exc()
            print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
            print(err_str)
        return True
    elif level == 'info':
        logging.info(message)
        return True
    else:
        return False


month_to_amount = {
    '1': 80,
    '3': 200,
    '6': 400,
}


def get_photo_id(photo_name: Literal[
    'start', 'support', 'referral', 'payment_type', 'my_subscribe', 'month_price', 'what_in_closed']):
    with open('config.json', 'r') as read_file:
        data = json.load(read_file)
        photo = data['photos'][photo_name]
    return photo


def update_photo_id(photo_name: Literal[
    'start', 'support', 'referral', 'payment_type', 'my_subscribe', 'month_price', 'what_in_closed'], new_id: str):
    with open('config.json', 'r') as read_file:
        data = json.load(read_file)

    with open('config.json', 'w') as write_file:
        data['photos'][photo_name] = new_id
        json.dump(data, write_file)
    return True


async def check_users():
    try:
        while True:
            try:
                await asyncio.sleep(24 * 60 * 60)
                today = datetime.datetime.now()

                all_subscribes = await PaymentCore.find_all()
                for subscribe in all_subscribes:
                    if subscribe.will_end_at < today:
                        user = await UserCore.find_one(id=subscribe.usertable_id)
                        for chat_id in config.CHATS_FOLDER_IDS:
                            await bot.ban_chat_member(
                                chat_id=chat_id,
                                user_id=user.id
                            )
                        await PaymentCore.delete(id=subscribe.id)
                        await bot.send_message(chat_id=user.user_id,
                                               text="<b>Ваша подписка окончена, вы исключены из чатов</b>")
                    elif (subscribe.will_end_at - today).days == 3:
                        user = await UserCore.find_one(id=subscribe.usertable_id)
                        await bot.send_message(chat_id=user.user_id,
                                               text="<b>До окончания вашей подписки осталось 3 дня</b>")

            except Exception as err:
                err_str = traceback.format_exc()
                print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
                print(err_str)
                await log(err_str)

    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


async def send_message(model: messages.MsgModel) -> Message:
    try:
        if not model.photo and not model.photo_name:
            sent = await bot.send_message(
                chat_id=model.id,
                text=model.text,
                reply_markup=InlineKeyboardMarkup(inline_keyboard=model.markup) if model.markup else model.markup,
                reply_to_message_id=model.reply_to_message_id,
                disable_notification=model.disable_notifications,
                disable_web_page_preview=model.disable_web_page_preview,
                parse_mode=model.parse_mode
            )
            await BotMessageCore.add(
                type='text',
                text=model.text,
                send_at=sent.date,
                message_id=sent.message_id
            )
            return sent
        else:
            photo = model.photo
            if model.photo_name:
                photo = get_photo_id(model.photo_name)
            if model.photo_type == 'filename':
                photo = FSInputFile(model.photo)
            start = True
            while True:
                try:
                    sent = await bot.send_photo(
                        chat_id=model.id,
                        caption=model.text,
                        photo=photo,
                        reply_markup=InlineKeyboardMarkup(
                            inline_keyboard=model.markup) if model.markup else model.markup,
                        reply_to_message_id=model.reply_to_message_id,
                        disable_notification=model.disable_notifications,
                        parse_mode=model.parse_mode
                    )
                    break
                except aiogram.exceptions.TelegramBadRequest:
                    if model.photo_name:
                        if start is False:
                            break
                        elif start is True:
                            start = False
                            photo = FSInputFile('photos/' + model.photo_name + '.png')
                    else:
                        return False
            if start is False:
                update_photo_id(model.photo_name, sent.photo[-1].file_id)

            await BotMessageCore.add(
                type='text',
                text=model.text,
                file_id=sent.photo[-1].file_id,
                send_at=sent.date,
                message_id=sent.message_id
            )
            return sent
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


async def delete_message(model: MsgModel) -> None:
    try:
        try:
            await bot.delete_message(chat_id=model.id, message_id=model.message_id)
        except aiogram.exceptions.TelegramBadRequest:
            pass
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


async def update_message(model: messages.MsgModel) -> None:
    try:
        try:
            if not model.photo and not model.photo_name:
                await bot.edit_message_text(
                    chat_id=model.id,
                    text=model.text,
                    message_id=model.message_id,
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=model.markup) if model.markup else model.markup,
                    disable_web_page_preview=model.disable_web_page_preview,
                    parse_mode=model.parse_mode
                )
            else:
                try:
                    await delete_message(model)
                except aiogram.exceptions.TelegramBadRequest:
                    pass
                await send_message(model)
        except aiogram.exceptions.TelegramBadRequest:
            try:
                await delete_message(model)
            except aiogram.exceptions.TelegramBadRequest:
                pass
            await send_message(model)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


async def save_user_message(message: Message, user_pk: int):
    try:
        content_type = message.content_type
        text = message.text or message.caption
        file_id = None
        if content_type == 'photo':
            file_id = message.photo[-1].file_id

        await UserMessageCore.add(
            usertable_id=user_pk,
            type=content_type,
            text=text,
            file_id=file_id,
            send_at=message.date,
            message_id=message.message_id
        )
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


class MsgTypeF(Filter):
    message_type = None

    def __init__(self, message_type: Literal['photo', 'text'] = None):
        """
        Check private or not
        and message content type photo or text
        :param message_type:
        """
        self.message_type = message_type

    async def __call__(self, update: Message, *args, **kwargs):
        check_message_type = update.content_type == self.message_type

        return check_message_type


class PrivateF(Filter):
    async def __call__(self, update: Union[Message, tg_types.CallbackQuery], *args, **kwargs):
        if type(update) == Message:
            chat = update.chat
        else:
            chat = update.message.chat

        return chat.type == 'private'


class ReFullmatchF(Filter):
    pattern = None

    def __init__(self, pattern: str):
        self.pattern = re.compile(pattern)

    async def __call__(self, callback: CallbackQuery, *args, **kwargs):
        rres = re.fullmatch(self.pattern, callback.data)
        if rres:
            return True
        else:
            return False


class NextMsgInfoCheckF(Filter):
    async def __call__(self, update: Message, *args, **kwargs):
        res = await UserCore.find_one(user_id=update.from_user.id)
        if res:
            next_msg_info = res.next_msg_info

            comparison_dict = {
                'rubles': 'photo',
                'crypto': 'text',
            }

            if update.text:
                with open('config.json') as read_f:
                    data = json.load(read_f)
                if data['enable_promo'] is True:
                    if update.text.upper().strip() in ('ATRUMS30', 'ATRUMS50'):
                        comparison_dict = {
                            'rubles': 'text',
                            'crypto': 'text',
                        }

            if next_msg_info:
                return comparison_dict[next_msg_info.split('/')[0]] == update.content_type
        else:
            return False


# next_message
@dp.message(PrivateF(), MsgTypeF('text'), NextMsgInfoCheckF())
async def start_update(message: Message):
    try:
        user = message.from_user
        text = message.text

        db_user = await UserCore.find_one(user_id=user.id)
        next_msg_info_split = db_user.next_msg_info.split('/')
        with open('config.json') as read_f:
            data = json.load(read_f)
        if data['enable_promo'] is True:
            if text.upper().strip() in ('ATRUMS30', 'ATRUMS50'):
                if next_msg_info_split[1].count(',') == 0:
                    month_count = next_msg_info_split[1] + (',0.3' if text.upper().strip() == 'ATRUMS30' else ',0.5')
                else:
                    month_count = next_msg_info_split[1]
                await UserCore.update(filter_by={'user_id': user.id},
                                      next_msg_info=f'{next_msg_info_split[0]}/{month_count}/{message.message_id}')
                if next_msg_info_split[0] == 'crypto':
                    msg = messages.PayAccessCrypto(id=user.id, message_id=message.message_id)
                    msg.text = msg.text.format(day_count=str(int(month_count.split(',')[0]) * 30),
                                               pay_amount=str(month_to_amount[month_count]))
                    await update_message(msg)
                return

        count_month = next_msg_info_split[1]
        right_amount = month_to_amount[count_month]

        async def send_error():
            await delete_message(MsgModel(id=user.id, message_id=int(db_user.next_msg_info.split('/')[2])))
            await send_message(messages.PayError(id=user.id))
            await UserCore.update(filter_by={'user_id': user.id}, next_msg_info=None)

        async with httpx.AsyncClient() as client:
            response = await client.get('https://apilist.tronscanapi.com/api/transaction-info',
                                        params={'hash': text.strip()})
            json_req = response.json()

        if json_req == {}:
            return await send_error()

        contract_ret = json_req['contractRet']
        if contract_ret != 'SUCCESS':
            return await send_error()

        confirmed = json_req['confirmed']
        if not confirmed is True:
            return await send_error()

        to_address = json_req['toAddress']
        if to_address != config.TO_ADDRESS:
            return await send_error()

        amount = float(json_req['trc20TransferInfo'][0]['amount_str']) / 1000000
        if amount != right_amount:
            return await send_error()

        await delete_message(MsgModel(id=user.id, message_id=int(db_user.next_msg_info.split('/')[2])))

        markup = []

        for i in config.CHATS_FOLDER_IDS:
            res = await bot.create_chat_invite_link(chat_id=i, member_limit=1)
            res2 = await bot.get_chat(chat_id=i)
            markup.append([IButton(text=res2.first_name, url=res.invite_link)])

        msg = messages.PaySuccess(id=user.id)
        msg.markup = markup
        await send_message(msg)

        will_end_at = datetime.datetime.utcnow() + datetime.timedelta(days=30 * int(count_month))
        await PaymentCore.add(
            usertable_id=db_user.id,
            will_end_at=will_end_at,
            duration_in_month=int(count_month),
            payment_method='crypto',
            transaction_hash=text,
            currency='usdt',
            price=right_amount
        )

        await UserCore.update(filter_by={'user_id': user.id}, next_msg_info=None)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


# next_message
@dp.message(PrivateF(), MsgTypeF('photo'), NextMsgInfoCheckF())
async def start_update(message: Message):
    try:
        user = message.from_user

        db_user = await UserCore.find_one(user_id=user.id)

        count_month = db_user.next_msg_info.split('/')[1]
        sent_message_id = db_user.next_msg_info.split('/')[2]

        await delete_message(MsgModel(id=user.id, message_id=sent_message_id))

        msg_to_user = messages.RublesAfterPayInfo(id=user.id)
        await send_message(msg_to_user)

        await bot.forward_message(config.ADMIN_ID, user.id, message.message_id)
        msg_to_admin = messages.AdminCheckCheck(id=config.ADMIN_ID)
        msg_to_admin.text = msg_to_admin.text.format(username=user.username or 'username', month_count=count_month)
        msg_to_admin.markup = [[
            IButton(text='Да', callback_data=f'{db_user.id}/{count_month}/yes'),
            IButton(text='Нет', callback_data=f'{db_user.id}/{count_month}/no'),
        ]]
        await send_message(msg_to_admin)

        await UserCore.update(filter_by={'user_id': user.id}, next_msg_info=None)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


# команда /start
@dp.message(PrivateF(), MsgTypeF('text'), F.text[:6] == '/start')
async def start_update(message: Message):
    try:
        raise "asgfsjhgkfdakdsfugiehwj"
        user = message.from_user
        text = message.text
        db_user = await UserCore.find_one(user_id=user.id)
        if not db_user:
            referral_id = None
            rres = re.fullmatch(r'/start ([0-9-]+)', text)
            if rres:
                referral_id = int(rres.group(1))

            await UserCore.add(
                user_id=user.id,
                username=user.username,
                first_name=user.first_name,
                referral_id=referral_id
            )
        db_user = await UserCore.find_one(user_id=user.id)

        await save_user_message(message, db_user.id)

        msg = messages.StartMessage(id=user.id)

        msg.text = msg.text.format(first_name=user.first_name or user.username)
        await send_message(msg)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


@dp.message(PrivateF(), MsgTypeF('text'), F.text[:6] == '/promo')
async def promo(message: Message):
    try:
        with open('config.json') as read_f:
            data = json.load(read_f)

        enable_promo = False if data['enable_promo'] is True else True

        with open('config.json', 'w') as write_f:
            json.dump({'enable_promo': enable_promo}, write_f)

        await send_message(MsgModel(id=message.from_user.id,
                                    text='Промокоды выключены' if enable_promo is False else 'Промокоды включены'))
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


# кнопка (отмена)
@dp.callback_query(PrivateF(), F.data == 'to_start')
async def to_start(callback: CallbackQuery):
    try:
        user = callback.from_user
        message = callback.message
        msg = messages.StartMessage(
            id=user.id,
            message_id=message.message_id
        )

        msg.text = msg.text.format(first_name=user.first_name or user.username)

        await update_message(msg)

        await UserCore.update(filter_by={'user_id': user.id}, next_msg_info=None)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


# кнопка (получить доступ)
@dp.callback_query(PrivateF(), F.data == 'get_access')
async def get_access1(callback: CallbackQuery):
    try:
        msg = messages.GetAccess(
            id=callback.from_user.id,
            message_id=callback.message.message_id
        )

        await update_message(msg)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


# кнопка выбора месяцев подписки
@dp.callback_query(PrivateF(), ReFullmatchF('get_access/(1|3|6)_month'))
async def get_access2(callback: CallbackQuery):
    try:
        cdata = callback.data
        user = callback.from_user
        message = callback.message
        rres = re.fullmatch(r'get_access/(1|3|6)_month', cdata)
        month_count = rres.group(1)

        msg = messages.GetAccessXMonth(
            id=user.id,
            message_id=message.message_id,
            markup=[
                [IButton(text='Криптовалюта', callback_data=f'get_access/{month_count}_month/crypto')],
                [IButton(text='Рубли', callback_data=f'get_access/{month_count}_month/rubles')]
            ]
        )

        await update_message(msg)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


# кнопка выбора метода оплаты
@dp.callback_query(PrivateF(), ReFullmatchF('get_access/(1|3|6)(,0.3|.0.5|)_month/(crypto|rubles)'))
async def get_access3(callback: CallbackQuery):
    try:
        cdata = callback.data
        user = callback.from_user
        message = callback.message

        rres = re.fullmatch(r'get_access/(1|3|6)(|,0.3|.0.5)_month/(crypto|rubles)', cdata)
        month_count = rres.group(1) + rres.group(2)
        pay_type = rres.group(3)

        if pay_type == 'crypto':
            async with httpx.AsyncClient() as client:
                response = await client.post("https://pay.crypt.bot/api/createInvoice",
                                             headers={
                                                 "Crypto-Pay-API-Token": config.CRYPTO_BOT_TOKEN},
                                             data={
                                                 "asset": "USDT",
                                                 "amount": month_to_amount[month_count]
                                             })

                pay_url = response.json()["result"]["pay_url"]
                invoice_id = response.json()["result"]["invoice_id"]

            msg = messages.PayAccessCrypto(id=user.id, message_id=message.message_id)
            msg.text = msg.text.format(day_count=str(int(month_count) * 30),
                                       pay_amount=str(month_to_amount[month_count]))
            msg.markup = [
                [IButton(text="Оплатить", url=pay_url)],
                [IButton(text="Оплатил",
                         callback_data=f"payed?invoice_id={invoice_id}&price={month_to_amount[month_count]}&month={month_count}")],
                [IButton(text='Отмена', callback_data='to_start')]
            ]

            await update_message(msg)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


@dp.callback_query(PrivateF(), ReFullmatchF('payed.+'))
async def payed(callback: CallbackQuery):
    try:
        cdata = callback.data
        user = callback.from_user
        message = callback.message

        invoice_id = int(re.search('invoice_id=([^?&/]+)', cdata).group(1))
        price = int(re.search('price=([^?&/]+)', cdata).group(1))
        month = int(re.search('month=([^?&/]+)', cdata).group(1))

        try:
            async with httpx.AsyncClient() as client:
                response = await client.post("https://pay.crypt.bot/api/getinvoices",
                                             headers={
                                                 "Crypto-Pay-API-Token": config.CRYPTO_BOT_TOKEN},
                                             data={
                                                 "invoice_ids": [invoice_id],
                                             })

            status = response.json()["result"]["items"][0]["status"]

            if status == "active":
                await bot.answer_callback_query(callback_query_id=callback.id, text="Вы ещё не произвели оплату",
                                                show_alert=True)
            elif status == "paid":
                db_user = await UserCore.find_one(user_id=user.id)

                markup = []

                for i in config.CHATS_FOLDER_IDS:
                    res = await bot.create_chat_invite_link(chat_id=i, member_limit=1)
                    res2 = await bot.get_chat(chat_id=i)
                    markup.append([IButton(text=res2.title or res2.first_name or "Ссылка", url=res.invite_link)])

                msg = messages.PaySuccess(id=user.id)
                msg.markup = markup
                await send_message(msg)
                await delete_message(MsgModel(id=user.id, message_id=message.message_id))
                will_end_at = datetime.datetime.utcnow() + datetime.timedelta(days=30 * month)
                await PaymentCore.add(
                    usertable_id=db_user.id,
                    will_end_at=will_end_at,
                    duration_in_month=month,
                    payment_method='crypto',
                    transaction_hash="",
                    currency='usdt',
                    price=price
                )
            else:
                raise
        except:
            traceback.print_exc()
            await bot.answer_callback_query(callback_query_id=callback.id, text="Что-то не так", show_alert=True)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


@dp.callback_query(PrivateF(), F.data == 'my_subscribe')
async def my_subscribe(callback: CallbackQuery):
    try:
        user = callback.from_user
        db_user = await UserCore.find_one(user_id=user.id)
        subscribe = await PaymentCore.find_one(usertable_id=db_user.id)

        if subscribe:
            msg = messages.MySubscribeActive(id=user.id, message_id=callback.message.message_id)
            residue = str((subscribe.will_end_at - datetime.datetime.utcnow()).days)
            msg.text = msg.text.format(residue=residue, username=user.username or 'username')
            await update_message(msg)
        else:
            await update_message(messages.MySubscribeInactive(id=user.id, message_id=callback.message.message_id))
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


@dp.callback_query(PrivateF(), F.data == 'support')
async def support(callback: CallbackQuery):
    try:
        user = callback.from_user
        await update_message(messages.Support(id=user.id, message_id=callback.message.message_id))
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


@dp.callback_query(PrivateF(), F.data == 'referral')
async def referral(callback: CallbackQuery):
    try:
        user = callback.from_user
        db_user = await UserCore.find_one(user_id=user.id)
        referrals = await UserCore.find_all(referral_id=db_user.id)

        referral_count = str(len(referrals))
        referral_url = f'https://t.me/{config.BOT_USERNAME}?start={db_user.id}'
        list_referrals = ''
        for ref in referrals:
            if ref.username:
                list_referrals += '\n@' + ref.username
                if ref.first_name:
                    list_referrals += '(' + ref.first_name + ')'
            elif ref.first_name:
                list_referrals += '\n' + ref.first_name
            else:
                list_referrals += '\nnoname'
        msg = messages.Referral(id=user.id, message_id=callback.message.message_id)
        msg.text = msg.text.format(balance=f'{db_user.balance:.2f}',
                                   referral_count=referral_count,
                                   referral_url=referral_url,
                                   referrals=list_referrals[:3000])
        await update_message(msg)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


@dp.callback_query(PrivateF(), F.data == 'more_info')
async def more_info(callback: CallbackQuery):
    try:
        user = callback.from_user
        await update_message(messages.MoreInfo(id=user.id, message_id=callback.message.message_id))
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


@dp.message(PrivateF(), MsgTypeF('text'))
async def table(message: Message):
    try:
        if message.from_user.id == config.ADMIN_ID:
            if message.text.lower().strip() in ("таблица", "table", "/table"):
                file_name = 'users_table.xlsx'

                wb = Workbook()
                ws = wb.active
                ws.title = 'Выгрузка базы данных'

                res = await PaymentCore.find_all()

                ws.column_dimensions['A'].width = 3
                ws.column_dimensions['B'].width = 13
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 40
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['E'].width = 10

                ws.cell(row=1, column=1, value='ID')
                ws.cell(row=1, column=2, value='USERNAME')
                ws.cell(row=1, column=3, value='СУММА ОПЛАТЫ')
                ws.cell(row=1, column=4, value='ДАТА ОПЛАТЫ')
                ws.cell(row=1, column=5, value='СРОК')
                ws.cell(row=1, column=6, value='ОПЛАТА ПО СЧËТУ')

                row = 2

                user_payment_count = {}

                all_payments_count = 0
                thirty_days_payments_count = 0
                month_payments_count = 0

                today = datetime.datetime.now()

                for i in res:
                    if user_payment_count.get(i.usertable_id):
                        user_payment_count[i.usertable_id] += 1
                    else:
                        user_payment_count[i.usertable_id] = 1

                    all_payments_count += 1
                    if (today - i.bought_at).days < 30:
                        thirty_days_payments_count += 1
                    if today.month == i.bought_at.month and today.year == i.bought_at.year:
                        month_payments_count += 1

                    user = await UserCore.find_one(id=i.usertable_id)
                    ws.cell(row=row, column=1, value=user.id)
                    ws.cell(row=row, column=2, value=user.username or "Без юзернейма")
                    ws.cell(row=row, column=3, value=i.price)
                    ws.cell(row=row, column=4, value=i.bought_at)
                    ws.cell(row=row, column=5, value=str(i.duration_in_month) + " мес.")
                    ws.cell(row=row, column=6, value=str(user_payment_count[i.usertable_id]))
                    row += 1

                wb.save(file_name)

                caption = "Всего оплат: <b>" + str(all_payments_count) + "</b>\nОплат за 30 дней: <b>" + str(
                    thirty_days_payments_count) + "</b>\nОплат за месяц: <b>" + str(month_payments_count) + "</b>"

                await bot.send_document(
                    chat_id=message.from_user.id,
                    document=FSInputFile(file_name),
                    caption=caption
                )
                os.remove(file_name)
    except Exception as err:
        err_str = traceback.format_exc()
        print(f"\033[1;31mERROR:\033[37m {err}\033[0m")
        print(err_str)
        await log(err_str)


allowed_updates = ['message', 'callback_query']


async def start_polling():
    while True:
        try:
            # noinspection PyAsyncCall
            await dp.start_polling(bot, polling_timeout=300, handle_signals=False, allowed_updates=allowed_updates)
        except:
            time.sleep(10)
            continue
